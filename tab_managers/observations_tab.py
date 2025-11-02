"""Observations tab manager for the observation log application."""

import os
import datetime
from PyQt6.QtWidgets import QWidget, QMessageBox, QTableWidgetItem, QFileDialog, QToolButton, QMenu
from PyQt6.QtCore import QStringListModel
from PyQt6.QtGui import QColor
from PyQt6 import uic

from dialogs import EditObservationDialog
from calculations import calculate_angular_separation
from utilities import NumericTableWidgetItem
import settings


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ObservationsTabManager:
    """Manages the Observations tab functionality."""
    
    def __init__(self, parent, db, tab_widget, statusbar):
        """
        Initialize the Observations tab manager.
        
        Args:
            parent: Parent window
            db: Database instance
            tab_widget: Main tab widget
            statusbar: Status bar for messages
        """
        self.parent = parent
        self.db = db
        self.tab_widget = tab_widget
        self.statusbar = statusbar
        self.setup_tab()
    
    def setup_tab(self):
        """Setup the Observations tab."""
        observation_widget = QWidget()
        base_dir = os.path.dirname(os.path.dirname(__file__))
        ui_path = os.path.join(base_dir, 'observation_tab.ui')
        uic.loadUi(ui_path, observation_widget)
        self.tab_widget.addTab(observation_widget, "Observations")
        
        # Store references
        self.observations_table = observation_widget.findChild(QWidget, "observationsTable")
        self.observation_filter_list_view = observation_widget.findChild(QWidget, "observationFilterListView")
        self.session_id_combo_box = observation_widget.findChild(QWidget, "sessionIdComboBox")
        self.object_combo_box = observation_widget.findChild(QWidget, "objectComboBox")
        self.camera_combo_box = observation_widget.findChild(QWidget, "cameraComboBox")
        self.telescope_combo_box = observation_widget.findChild(QWidget, "telescopeComboBox")
        self.filter_combo_box = observation_widget.findChild(QWidget, "filterComboBox")
        self.image_count_spin_box = observation_widget.findChild(QWidget, "imageCountSpinBox")
        self.exposure_length_spin_box = observation_widget.findChild(QWidget, "exposureLengthSpinBox")
        self.comments_line_edit = observation_widget.findChild(QWidget, "commentsLineEdit")
        self.add_observation_button = observation_widget.findChild(QWidget, "addObservationButton")
        self.edit_observation_button = observation_widget.findChild(QWidget, "editObservationButton")
        self.export_observation_button = observation_widget.findChild(QWidget, "exportObservationButton")
        self.delete_observation_button = observation_widget.findChild(QWidget, "deleteObservationButton")

        # Setup filter list view
        self.filter_model = QStringListModel()
        self.observation_filter_list_view.setModel(self.filter_model)
        
        # Setup export button with popup menu
        self.export_menu = QMenu(self.export_observation_button)
        self.export_action = self.export_menu.addAction("Export to Excel")
        self.export_action.triggered.connect(self.export_observations_to_excel)
        self.export_html_action = self.export_menu.addAction("Export to HTML")
        self.export_html_action.triggered.connect(self.export_observations_to_html)
        self.export_observation_button.setMenu(self.export_menu)
        
        # Connect other signals
        self.add_observation_button.clicked.connect(self.add_observation)
        self.edit_observation_button.clicked.connect(self.edit_observation)
        self.delete_observation_button.clicked.connect(self.delete_observation)
        self.observation_filter_list_view.selectionModel().currentChanged.connect(self.filter_observations)
        
        # Hide ID column
        self.observations_table.setColumnHidden(0, True)
        self.observations_table.setColumnWidth(1, 100)  # Session ID
        self.observations_table.setColumnWidth(2, 80)   # Date
        self.observations_table.setColumnWidth(3, 100)  # Object
        self.observations_table.setColumnWidth(4, 100)  # Camera
        self.observations_table.setColumnWidth(5, 100)  # Telescope
        self.observations_table.setColumnWidth(6, 80)   # Filter
        self.observations_table.setColumnWidth(7, 60)   # Images
        self.observations_table.setColumnWidth(8, 80)   # Exposure
        self.observations_table.setColumnWidth(9, 100)  # Total Exposure
        self.observations_table.setColumnWidth(10, 80)  # Moon Illumination
        self.observations_table.setColumnWidth(11, 100) # Angular Separation
        self.observations_table.setColumnWidth(12, 200) # Comments
        
        self.load_observations()
        self.update_observation_combos()
        self.update_filter_list()
    
    def update_observation_combos(self):
        """Update all combo boxes in the observations tab."""
        try:
            # Update Session ID combo
            sessions = self.db.get_all_sessions()
            self.session_id_combo_box.clear()
            for session in sessions:
                self.session_id_combo_box.addItem(session['session_id'])

            # Update Object combo
            objects = self.db.get_all_objects()
            self.object_combo_box.clear()
            for obj in objects:
                self.object_combo_box.addItem(obj['name'])

            # Update Camera combo
            cameras = self.db.get_all_cameras()
            self.camera_combo_box.clear()
            for camera in cameras:
                self.camera_combo_box.addItem(camera['name'])

            # Update Telescope combo
            telescopes = self.db.get_all_telescopes()
            self.telescope_combo_box.clear()
            for telescope in telescopes:
                self.telescope_combo_box.addItem(telescope['name'])

            # Update Filter combo
            filters = self.db.get_all_filters()
            self.filter_combo_box.clear()
            for filt in filters:
                self.filter_combo_box.addItem(filt['name'])
        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to update observation combos: {str(e)}')

    def update_filter_list(self):
        """Update the filter list view with unique object names from observations."""
        try:
            observations = self.db.get_all_observations()
            unique_objects = set(obs['object_name'] for obs in observations)
            filter_items = ['< All Names >'] + sorted(list(unique_objects))
            self.filter_model.setStringList(filter_items)
            # Set default selection to "< All Names >"
            self.observation_filter_list_view.setCurrentIndex(self.filter_model.index(0, 0))
        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to update filter list: {str(e)}')
    
    def load_observations(self, object_filter=None):
        """Load observations from database and display in table, optionally filtered by object name."""
        try:
            if object_filter == '< All Names >' or object_filter is None:
                observations = self.db.get_all_observations()
            else:
                observations = [obs for obs in self.db.get_all_observations() if obs['object_name'] == object_filter]

            self.observations_table.setRowCount(len(observations))

            for row, obs in enumerate(observations):
                # Calculate angular separation if coordinates are available
                angular_sep = ""
                angular_sep_value = None
                if obs['ra'] is not None and obs['dec'] is not None and obs['moon_ra'] is not None and obs['moon_dec'] is not None:
                    try:
                        angular_sep_value = calculate_angular_separation(obs['ra']*15.0, obs['dec'], obs['moon_ra'], obs['moon_dec'])
                        angular_sep = f"{angular_sep_value:.0f}°"
                    except Exception:
                        angular_sep = "N/A"

                self.observations_table.setItem(row, 0, QTableWidgetItem(str(obs['id'])))
                self.observations_table.setItem(row, 1, QTableWidgetItem(obs['session_id']))
                self.observations_table.setItem(row, 2, QTableWidgetItem(obs['start_date']))
                self.observations_table.setItem(row, 3, QTableWidgetItem(obs['object_name']))
                self.observations_table.setItem(row, 4, QTableWidgetItem(obs['camera_name']))
                self.observations_table.setItem(row, 5, QTableWidgetItem(obs['telescope_name']))
                self.observations_table.setItem(row, 6, QTableWidgetItem(obs['filter_name']))
                self.observations_table.setItem(row, 7, NumericTableWidgetItem(obs['image_count']))
                self.observations_table.setItem(row, 8, NumericTableWidgetItem(obs['exposure_length']))
                self.observations_table.setItem(row, 9, NumericTableWidgetItem(obs['total_exposure']))

                # Moon Illumination column with conditional highlighting
                moon_illumination_item = QTableWidgetItem(f"{obs['moon_illumination']:.0f}%" if obs['moon_illumination'] is not None else "")
                moon_illumination_warning = settings.get_moon_illumination_warning()
                if obs['moon_illumination'] is not None and obs['moon_illumination'] > moon_illumination_warning:
                    # For pastel: high value, low-to-medium saturation
                    saturation = 40  # 0-100, lower = more pastel
                    value_hsv = 95   # 0-100, brightness
                    color = QColor.fromHsv(0, int(saturation * 255 / 100), int(value_hsv * 255 / 100))
                    moon_illumination_item.setBackground(color)
                self.observations_table.setItem(row, 10, moon_illumination_item)

                # Angular Separation column with conditional highlighting
                angular_sep_item = QTableWidgetItem(angular_sep)
                angular_sep_warning = settings.get_moon_angular_separation_warning()
                if angular_sep_value is not None and angular_sep_value < angular_sep_warning:
                    # For pastel: high value, low-to-medium saturation
                    saturation = 40  # 0-100, lower = more pastel
                    value_hsv = 95   # 0-100, brightness
                    color = QColor.fromHsv(0, int(saturation * 255 / 100), int(value_hsv * 255 / 100))
                    angular_sep_item.setBackground(color)
                self.observations_table.setItem(row, 11, angular_sep_item)

                self.observations_table.setItem(row, 12, QTableWidgetItem(obs['comments'] or ''))
            self.observations_table.resizeColumnsToContents()
            self.statusbar.showMessage(f'Loaded {len(observations)} observation(s)')
        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to load observations: {str(e)}')
    
    def add_observation(self):
        """Add a new observation to the database."""
        session_id = self.session_id_combo_box.currentText()
        object_name = self.object_combo_box.currentText()
        camera_name = self.camera_combo_box.currentText()
        telescope_name = self.telescope_combo_box.currentText()
        filter_name = self.filter_combo_box.currentText()
        image_count = self.image_count_spin_box.value()
        exposure_length = self.exposure_length_spin_box.value()
        comments = self.comments_line_edit.text().strip()
        
        if not all([session_id, object_name, camera_name, telescope_name, filter_name]):
            QMessageBox.warning(self.parent, 'Warning', 'Please select all required fields.')
            return
        
        if image_count == 0 or exposure_length == 0:
            QMessageBox.warning(
                self.parent, 'Warning',
                'Please enter valid image count and exposure length.'
            )
            return
        
        try:
            self.db.add_observation(session_id, object_name, camera_name, telescope_name,
                                    filter_name, image_count, exposure_length, comments)
            self.image_count_spin_box.setValue(0)
            self.exposure_length_spin_box.setValue(0)
            self.comments_line_edit.clear()
            # Reload with current filter
            current_filter = self.get_current_filter()
            self.load_observations(current_filter)
            self.update_filter_list()  # Update filter list in case new object was added
            self.statusbar.showMessage(f'Added observation for {object_name}')
        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to add observation: {str(e)}')
    
    def edit_observation(self):
        """Edit the selected observation."""
        selected_rows = self.observations_table.selectedItems()
        
        if not selected_rows:
            QMessageBox.warning(self.parent, 'Warning', 'Please select an observation to edit.')
            return
        
        row = self.observations_table.currentRow()
        observation_id = int(self.observations_table.item(row, 0).text())
        current_session_id = self.observations_table.item(row, 1).text()
        current_object = self.observations_table.item(row, 3).text()
        current_camera = self.observations_table.item(row, 4).text()
        current_telescope = self.observations_table.item(row, 5).text()
        current_filter = self.observations_table.item(row, 6).text()
        current_image_count = int(self.observations_table.item(row, 7).text())
        current_exposure = int(self.observations_table.item(row, 8).text())
        current_comments = self.observations_table.item(row, 12).text()
        
        # Get available options
        sessions = [s['session_id'] for s in self.db.get_all_sessions()]
        objects = [o['name'] for o in self.db.get_all_objects()]
        cameras = [c['name'] for c in self.db.get_all_cameras()]
        telescopes = [t['name'] for t in self.db.get_all_telescopes()]
        filters = [f['name'] for f in self.db.get_all_filters()]
        
        dialog = EditObservationDialog(
            current_session_id, current_object, current_camera, current_telescope,
            current_filter, current_image_count, current_exposure, current_comments,
            sessions, objects, cameras, telescopes, filters, self.parent
        )
        
        if dialog.exec():
            session_id, object_name, camera_name, telescope_name, filter_name, \
                image_count, exposure_length, comments = dialog.get_values()
            try:
                self.db.update_observation(observation_id, session_id, object_name,
                                           camera_name, telescope_name, filter_name,
                                           image_count, exposure_length, comments)
                # Reload with current filter
                current_filter = self.get_current_filter()
                self.load_observations(current_filter)
                self.update_filter_list()  # Update filter list in case object name changed
                self.statusbar.showMessage(f'Updated observation ID {observation_id}')
            except Exception as e:
                QMessageBox.critical(self.parent, 'Error', f'Failed to update observation: {str(e)}')
    
    def delete_observation(self):
        """Delete the selected observation."""
        selected_rows = self.observations_table.selectedItems()

        if not selected_rows:
            QMessageBox.warning(self.parent, 'Warning', 'Please select an observation to delete.')
            return

        row = self.observations_table.currentRow()
        observation_id = int(self.observations_table.item(row, 0).text())
        object_name = self.observations_table.item(row, 3).text()
        moon_illumination = self.observations_table.item(row, 10).text()
        angular_sep = self.observations_table.item(row, 11).text()

        reply = QMessageBox.question(
            self.parent, 'Confirm Deletion',
            f'Are you sure you want to delete observation for "{object_name}"?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.delete_observation(observation_id)
                # Reload with current filter
                current_filter = self.get_current_filter()
                self.load_observations(current_filter)
                self.statusbar.showMessage(f'Deleted observation for {object_name}')
            except Exception as e:
                QMessageBox.critical(self.parent, 'Error', f'Failed to delete observation: {str(e)}')

    def filter_observations(self, current, previous):
        """Filter observations based on selected object name."""
        selected_index = current
        if selected_index.isValid():
            selected_text = self.filter_model.data(selected_index, 0)
            self.load_observations(selected_text)

    def get_current_filter(self):
        """Get the currently selected filter."""
        current_index = self.observation_filter_list_view.currentIndex()
        if current_index.isValid():
            return self.filter_model.data(current_index, 0)
        return '< All Names >'

    def export_observations_to_excel(self):
        """Export observations to Excel file with same columns as visible in the table."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self.parent, 'Error',
                               'openpyxl is required for Excel export. Please install it with: pip install openpyxl')
            return
        
        try:
            # Check if openpyxl is available
            if not OPENPYXL_AVAILABLE:
                QMessageBox.critical(self.parent, 'Error',
                                   'openpyxl library not found. Please install it first.')
                return

            # Get current observations with same filtering as table
            current_filter = self.get_current_filter()
            if current_filter == '< All Names >' or current_filter is None:
                observations = self.db.get_all_observations()
            else:
                observations = [obs for obs in self.db.get_all_observations() if obs['object_name'] == current_filter]

            if not observations:
                QMessageBox.information(self.parent, 'Information', 'No observations to export.')
                return

            # Create a proper filename based on current date and filter
            current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if current_filter and current_filter != '< All Names >':
                safe_filter = "".join(c for c in current_filter if c.isalnum() or c in (' ', '-', '_')).rstrip()
                default_filename = f"observations_{safe_filter}_{current_date}.xlsx"
            else:
                default_filename = f"observations_all_{current_date}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self.parent,
                'Save Observations Excel File',
                default_filename,
                'Excel Files (*.xlsx)'
            )

            if not file_path:
                return

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            wb.iso_dates = True
            ws = wb.active
            ws.title = "Observations"

            # Define headers in same order as table columns (excluding ID)
            headers = [
                'Session ID', 'Date', 'Object', 'Camera', 'Telescope', 'Filter',
                'Images', 'Exposure (s)', 'Total Exposure (s)', 'Moon Illumination (%)',
                'Angular Separation', 'Comments'
            ]

            # Write headers with formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            alignment = Alignment(horizontal='center')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Write data rows
            for row, obs in enumerate(observations, 2):
                # Calculate angular separation same way as in load_observations
                angular_sep = ""
                angular_sep_value = None
                if obs['ra'] is not None and obs['dec'] is not None and obs['moon_ra'] is not None and obs['moon_dec'] is not None:
                    try:
                        angular_sep_value = calculate_angular_separation(obs['ra']*15.0, obs['dec'], obs['moon_ra'], obs['moon_dec'])
                    except Exception:
                        angular_sep_value = ""

                # Write row data in same order as headers
                row_data = [
                    obs['session_id'],
                    datetime.datetime.strptime(obs['start_date'],"%Y-%m-%d"),
                    obs['object_name'],
                    obs['camera_name'],
                    obs['telescope_name'],
                    obs['filter_name'],
                    obs['image_count'],
                    obs['exposure_length'],
                    obs['total_exposure'],
                    float(obs['moon_illumination'])/100.0 if obs['moon_illumination'] is not None else "",
                    angular_sep_value,
                    obs['comments'] or ''
                ]

                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'  # Date format
                ws.cell(row=row, column=10).number_format = '0%'  # Moon Illumination as percentage
                ws.cell(row=row, column=11).number_format = '0°'    # Angular Separation in degrees

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(file_path)
            self.statusbar.showMessage(f'Observations exported to {file_path}')
            QMessageBox.information(self.parent, 'Success', f'Observations exported successfully to:\n{file_path}')

        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to export observations: {str(e)}')

    def export_observations_to_html(self):
        """Export observations to HTML file with same columns as visible in the table."""
        try:
            # Get current observations with same filtering as table
            current_filter = self.get_current_filter()
            if current_filter == '< All Names >' or current_filter is None:
                observations = self.db.get_all_observations()
            else:
                observations = [obs for obs in self.db.get_all_observations() if obs['object_name'] == current_filter]

            if not observations:
                QMessageBox.information(self.parent, 'Information', 'No observations to export.')
                return

            # Create a proper filename based on current date and filter
            current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if current_filter and current_filter != '< All Names >':
                safe_filter = "".join(c for c in current_filter if c.isalnum() or c in (' ', '-', '_')).rstrip()
                default_filename = f"observations_{safe_filter}_{current_date}.html"
            else:
                default_filename = f"observations_all_{current_date}.html"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self.parent,
                'Save Observations HTML File',
                default_filename,
                'HTML Files (*.html)'
            )

            if not file_path:
                return

            # Load HTML template
            base_dir = os.path.dirname(os.path.dirname(__file__))
            template_path = os.path.join(base_dir, 'templates', 'observations_export.html')
            
            try:
                with open(template_path, 'r', encoding='utf-8') as f:
                    template_content = f.read()
            except FileNotFoundError:
                QMessageBox.critical(self.parent, 'Error',
                                   'HTML template file not found. Please ensure templates/observations_export.html exists.')
                return

            # Generate table rows
            table_rows = ""
            for obs in observations:
                # Calculate angular separation same way as in load_observations
                angular_sep = ""
                angular_sep_value = None
                moon_warning_class = ""
                angular_warning_class = ""
                
                if obs['ra'] is not None and obs['dec'] is not None and obs['moon_ra'] is not None and obs['moon_dec'] is not None:
                    try:
                        angular_sep_value = calculate_angular_separation(obs['ra']*15.0, obs['dec'], obs['moon_ra'], obs['moon_dec'])
                        angular_sep = f"{angular_sep_value:.0f}°"
                    except Exception:
                        angular_sep = "N/A"

                # Check for moon illumination warning
                moon_illumination_warning = settings.get_moon_illumination_warning()
                if obs['moon_illumination'] is not None and obs['moon_illumination'] > moon_illumination_warning:
                    moon_warning_class = "moon-warning"

                # Check for angular separation warning
                angular_sep_warning = settings.get_moon_angular_separation_warning()
                if angular_sep_value is not None and angular_sep_value < angular_sep_warning:
                    angular_warning_class = "angular-warning"

                table_rows += f"""                <tr>
                    <td>{obs['session_id']}</td>
                    <td>{obs['start_date']}</td>
                    <td>{obs['object_name']}</td>
                    <td>{obs['camera_name']}</td>
                    <td>{obs['telescope_name']}</td>
                    <td>{obs['filter_name']}</td>
                    <td style="text-align: center;">{obs['image_count']}</td>
                    <td style="text-align: center;">{obs['exposure_length']}</td>
                    <td style="text-align: center;">{obs['total_exposure']}</td>
                    <td class="{moon_warning_class}" style="text-align: center;">{f"{obs['moon_illumination']:.0f}%" if obs['moon_illumination'] is not None else ""}</td>
                    <td class="{angular_warning_class}" style="text-align: center;">{angular_sep}</td>
                    <td>{obs['comments'] or ''}</td>
                </tr>
"""

            # Replace placeholders in template
            filter_name = current_filter if current_filter and current_filter != '< All Names >' else 'All Objects'
            export_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            completion_date = datetime.datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
            
            html_content = template_content.replace('{{filter_name}}', filter_name)
            html_content = html_content.replace('{{export_date}}', export_date)
            html_content = html_content.replace('{{total_records}}', str(len(observations)))
            html_content = html_content.replace('{{table_rows}}', table_rows)
            html_content = html_content.replace('{{completion_date}}', completion_date)

            # Write HTML file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.statusbar.showMessage(f'Observations exported to {file_path}')
            QMessageBox.information(self.parent, 'Success', f'Observations exported successfully to:\n{file_path}')

        except Exception as e:
            QMessageBox.critical(self.parent, 'Error', f'Failed to export observations: {str(e)}')